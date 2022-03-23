from random import randint
import openpyxl
import xlrd
from xlwt import Workbook
from xlutils.copy import copy

print("-----Bank Management-----")
print(" 1. New Account ")
print(" 2. Deposit Amount ")
print(" 3. Withdraw Amount ")
print(" 4. Balance Enquiry ")
print(" 5. Account Holder's Details ")
print(" 6. Close An Account ")
choice = int(input("Select your option: "))

new_account = []
deposit_amount = 0
withdraw_amount = []
balance_enquiry = []
acc_holders_details = []
acc_close = []

should_continue = True

wb = Workbook()
sheet1 = wb.add_sheet("Sheet 1", cell_overwrite_ok=True)
sheet1.write(0, 0, "Account Holder's Name")
sheet1.write(0, 1, "Mode Of Account")
sheet1.write(0, 2, "Balance Amount")
sheet1.write(0, 3, "Pin Number")


def add_new_user():
	suff_balance = True

	name = input("Enter The Account Holder's Name: ")
	mode = input("Enter the mode of account [C/S]:").upper()
	init_amnt = int(input("The Initial amount(>=500 for Saving and >=1000 for Current):"))
	while suff_balance:
		if mode == "C" and init_amnt < 1000 or mode == "S" and init_amnt < 500:
			init_amnt = int(
				input("The Initial amount(>=500 for Saving and >=1000 for Current, Please re-enter the amount: ):"))
		else:
			suff_balance = False
	pin = randint(1000, 9999)
	print("Your Pin Number Is: ", pin)
	new_account.append([name, mode, init_amnt, pin])
	for i in range(len(new_account)):
		for j in range(len(new_account[i])):
			sheet1.write(i + 1, 0, new_account[i][0])
			sheet1.write(i + 1, 1, new_account[i][1])
			sheet1.write(i + 1, 2, new_account[i][2])
			sheet1.write(i + 1, 3, new_account[i][3])
			wb.save("Account.xlsx")

	return pin, init_amnt


def add_deposit():
	dep_amnt = int(input("Enter The Deposit amount: "))
	user_pin = int(input("Enter Your Pin Number: "))

	loc = "Account.xlsx"

	rb = xlrd.open_workbook(loc, formatting_info=True)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)

	for row_idx in range(1, r_sheet.nrows):
		check_row = int(r_sheet.cell(row_idx, 3).value)
		if check_row == user_pin:
			init_val = int(r_sheet.cell(row_idx, 2).value)
			print(init_val, init_val + dep_amnt)
			w_sheet.write(row_idx, 2, init_val + dep_amnt)
	wb.save("Account.xlsx")


def withdraw():
	with_amnt = int(input("Enter The Amount To Be Withdrawn: "))
	user_pin = int(input("Enter Your Pin Number: "))

	loc = "Account.xlsx"

	rb = xlrd.open_workbook(loc, formatting_info=True)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)
	w_sheet = wb.get_sheet(0)

	for row_idx in range(1, r_sheet.nrows):
		check_row = int(r_sheet.cell(row_idx, 3).value)
		if check_row == user_pin:
			init_val = int(r_sheet.cell(row_idx, 2).value)
			print(init_val, init_val - with_amnt)
			w_sheet.write(row_idx, 2, init_val - with_amnt)
	wb.save("Account.xlsx")


def balance():
	user_pin = int(input("Enter Your Pin Number: "))
	loc = "Account.xlsx"
	rb = xlrd.open_workbook(loc)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)

	for row_idx in range(1, r_sheet.nrows):
		check_row = int(r_sheet.cell(row_idx, 3).value)
		if check_row == user_pin:
			init_val = int(r_sheet.cell(row_idx, 2).value)
			print("Your Balance is: ", init_val)


def account_holder():
	user_pin = int(input("Enter Your Pin Number: "))
	loc = "Account.xlsx"
	rb = xlrd.open_workbook(loc)
	r_sheet = rb.sheet_by_index(0)
	wb = copy(rb)

	for row_idx in range(1, r_sheet.nrows):
		check_row = int(r_sheet.cell(row_idx, 3).value)
		if check_row == user_pin:
			user_name = str(r_sheet.cell(row_idx, 0).value)
			user_mode = str(r_sheet.cell(row_idx, 1).value)
			print("The Account Holder Is: ", user_name)
			print("The Account Mode Is: ", user_mode)


def delete_account():
	user_pin = int(input("Enter Your Pin Number: "))
	path = 'Account.xlsx'
	book = openpyxl.load_workbook(path)

	sheet = book['Sheet 1']

	for row_idx in range(1, sheet.max_row+1):
		check_row = sheet.cell(row=row_idx, column=3).value
		if check_row == user_pin:
			sheet.delete_rows(row_idx, 1)


if choice == 1:
	add_new_user()
	while should_continue:
		new_choice = input("Do you want to add another user: ").lower()
		if new_choice == "y":
			add_new_user()
		else:
			should_continue = False

elif choice == 2:
	add_deposit()

elif choice == 3:
	withdraw()

elif choice == 4:
	balance()

elif choice == 5:
	account_holder()

elif choice == 6:
	delete_account()
